import numpy as np
import sys
import openpyxl as pyxl
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.animation as animation
from matplotlib.figure import Figure
import PyQt5.QtWidgets as QtWidgets
import PyQt5.QtCore as QtCore
from scipy.interpolate import RegularGridInterpolator
from scipy.interpolate import griddata
from matplotlib import cm
import matplotlib.colors as colors
from matplotlib.backends.backend_qt5agg import (FigureCanvasQTAgg, NavigationToolbar2QT as NavigationToolbar)
from qtrangeslider import QRangeSlider
import csv


axisName={'x':'x','y':'y','z':'z','w':'w'}
voxel3D=False
interpDefaultStep=[0.5,0.5,0.5] #Step for interpolation function : [xStep,yStep,zStep]
animParameters={'interval':1000,'repeat':False,'repeat_delay':0,'fps':1,'writer':'imagemagick'}
colorMap=cm.PiYG_r

class GetData(QtWidgets.QPushButton):
    """
    Extract data in exel or tsv files
    return : x,y,z,w
    """
    def __init__(self,name):
        super().__init__(name)
        self.data=[]
        self.rowNumber=0

    def loadFile(self):
        loadDialog=QtWidgets.QFileDialog()
        pathFile=loadDialog.getOpenFileName(None,'Select Directory',"",'*.xlsx *.csv *.tsv')
        if(pathFile[0]!=''):
            ext=pathFile[0][-3:]
            if(ext=='tsv'):
                self.data,self.rowNumber=self.getDataTSV(pathFile[0])
            elif(ext=='lsx'):
                self.data,self.rowNumber=self.getDataExc(pathFile[0])
        return self.data,self.rowNumber

    def getDataExc(self,pathFile):
        fileExcel = pyxl.load_workbook(pathFile)
        sheetExcel=fileExcel['Feuil1']
        maxCol=sheetExcel.max_column
        maxRow=sheetExcel.max_row
        coord=np.empty((maxCol,maxRow)) #x,y,z,intensity
        indCol=0
        for col in sheetExcel.columns:
            indRow=0
            for cell in col:
                coord[indCol][indRow]=cell.value
                indRow+=1
            indCol+=1
        return coord[:4],maxRow

    def getDataTSV(self,pathFile):
        dataFile=csv.reader(open(pathFile),delimiter='\t')
        x,y,z,w=[],[],[],[]
        countLine=0
        for iLine in dataFile:
            try:
                float(iLine[0])
                x.append(iLine[0])
                y.append(iLine[1])
                z.append(iLine[2])
                w.append(iLine[3])
                countLine+=1
            except ValueError:
                axisName['x']=iLine[0]
                axisName['y']=iLine[1]
                axisName['z']=iLine[2]
                axisName['w']=iLine[3]
                countLine-=1


        return np.array((x,y,z,w),dtype=float),countLine


class Visu(FigureCanvasQTAgg):
    '''
    Class inherithed by all display figure class
    '''
    def __init__(self,fig):
        self.fig=fig
        super(Visu, self).__init__(self.fig)
        self.rowNumber=0
        self.data=[0]
        self.dataInterp=[0]
        self.xMin,self.yMin,self.zMin,self.wMin=0,0,0,0
        self.xMax,self.yMax,self.zMax,self.wMax=0,0,0,0
        self.xRange,self.yRange,self.zRange=0,0,0
        self.wAvr=0
        self.interpStepX, self.interpStepY, self.interpStepZ=interpDefaultStep[0],interpDefaultStep[1],interpDefaultStep[2]
        self.xMinStep,self.yMinStep,self.zMinStep=1,1,1
        self.wNorm,self.opacity=[],[]

    def plotScatterData(self,ax,x,y,z,value,alpha,**kwargs):
        '''
        value : define color of the point after the rule of the colormap
        alpha : set opacity of the point
        '''

        color=colorMap(value,alpha)
        ax.scatter(x,y,z,c=color,linewidths=5,depthshade=False)
        self.draw()

    def dataProcess(self,values,n):
        '''
        Compute the average of values, and normalize values thanks its maximum
        '''
        dProcess=np.copy(values)
        self.average(values,n) #Change average
        for i in range(dProcess.shape[0]):
            dProcess[i]=dProcess[i]/self.wMax
        return dProcess

    def gradOpacity(self,values,min,lowThreshold,highThreshold,max):
        '''
       Set opacity increasing linear at highThreshold to max and at lowThreshold to min. Betwen low and high opacity is zero
        '''
        opacity=np.copy(values)
        i=0
        for v in values:
            if(max>=v>=highThreshold):
                v=(v-highThreshold)/(max-highThreshold)
                opacity[i]=v
            elif(min<=v<=lowThreshold):
                v=(lowThreshold-v)/(lowThreshold-min)
                opacity[i]=v
            else:
                opacity[i]=0
            i+=1
        return opacity

    def newExtremum(self,x,y,z,w):
        '''
        Set extremum, set range, for all axis and value column
        '''
        self.xMin,self.xMax=self.searchExtremum(x)
        self.xMin=int(self.xMin)
        self.xMax=int(self.xMax)
        self.yMin,self.yMax=self.searchExtremum(y)
        self.yMin=int(self.yMin)
        self.yMax=int(self.yMax)
        self.zMin,self.zMax=self.searchExtremum(z)
        self.zMin=int(self.zMin)
        self.zMax=int(self.zMax)
        self.wMin,self.wMax=self.searchExtremum(w)
        self.xMinStep=self.step(x,self.interpStepX)
        self.yMinStep=self.step(y,self.interpStepY)
        self.zMinStep=self.step(z,self.interpStepZ)
        self.xRange=int((self.xMax-self.xMin)/self.xMinStep+1)
        self.yRange=int((self.yMax-self.yMin)/self.yMinStep+1)
        self.zRange=int((self.zMax-self.zMin)/self.zMinStep+1)

    def searchExtremum(self,array):
        maxValue,minValue=array[0],array[0]
        for v in array:
            if(maxValue<v):
                maxValue=v
            if(minValue>v):
                minValue=v
        return minValue,maxValue

    def step(self,axis,stepComp):
        '''
        Search the min step present in axis
        type(axis)=np.array
        '''
        minStep=self.xMax
        for i in range(1,len(axis)):
            if(minStep>(axis[i]-axis[i-1])>0):
                minStep=axis[i]-axis[i-1]
        if(minStep<stepComp):
            stepComp=minStep
        return minStep

    def average(self,array,n):
        average=0
        for v in array:
            average+=v
        self.wAvr=average/n

    def interpFunc(self,x,y,z,value,n):
        '''
        x,y,z : original points where value is knew
        n : max of the axis
        return :
        f : all value interpolated
        grid_*: 3D arrays with coord for values interpolated
        use scipy module
        '''
        P=np.zeros((n,3),dtype=float)
        for i in range(n):
            P[i][:]=(x[i],y[i],z[i])
        grid_x, grid_y, grid_z = np.mgrid[self.xMin:self.xMax+self.interpStepX:self.interpStepX, self.yMin:self.yMax+self.interpStepY:self.interpStepY, self.zMin:self.zMax+self.interpStepZ:self.interpStepZ]
        f=griddata(P,value,(grid_x, grid_y, grid_z),method='linear')
        return f,grid_x,grid_y,grid_z

    def mat3D(self,x,y,z,w,n1,n2,n3):
        '''
        convert 1 list of value (w) with its coordinates (x,y,z) to an matrix (array)

        '''
        mat=np.zeros((n1,n2,n3))
        index=0
        for i in range(n1):
            for j in range(n2):
                for k in range(n3):
                    if(index<len(w)):
                        if((i*self.xMinStep+self.xMin)==x[index] and (j*self.yMinStep+self.yMin)==y[index] and (k*self.zMinStep+self.zMin)==z[index]):
                            value=w[index]
                            index+=1
                        else:
                            value=self.wAvr
                    else:
                        value=self.wAvr
                    mat[i][j][k]=value
        return mat

    def arrayToList(self,array):
        '''
        Convert 3D array to a list
        '''
        L=[]
        for i in range(len(array[0])):
            for j in range(len(array[1])):
                for k in range(len(array[2])):
                    L.append(array[i][j][k])
        return L

    def copy(self,visu):
        '''
        copy attributes related to figure
        '''
        self.data=visu.data
        self.xMin=visu.xMin
        self.yMin=visu.yMin
        self.zMin=visu.zMin
        self.wMin=visu.wMin
        self.xMax=visu.xMax
        self.yMax=visu.yMax
        self.zMax=visu.zMax
        self.wMax=visu.wMax
        self.xRange,self.yRange,self.zRange=visu.xRange,visu.yRange,visu.zRange
        self.xMinStep,self.yMinStep,self.zMinStep=visu.xMinStep,visu.yMinStep,visu.zMinStep
        self.wNorm=visu.wNorm
        self.opacity=visu.opacity
        self.wAvr=visu.wAvr
        self.rowNumber=visu.rowNumber
        self.interpStepX, self.interpStepY, self.interpStepZ=visu.interpStepX, visu.interpStepY, visu.interpStepZ


class Visu3D(Visu):
    '''
    Sub-class of Visu, for 3D plot
    '''
    def __init__(self, parent=None, width=500, height=400, dpi=100):
        fig=Figure(figsize=(width, height), dpi=dpi)
        super().__init__(fig)
        self.ax=self.fig.subplots(subplot_kw={'projection': '3d'})

    def process3D(self,data,rowNumber):
        self.data=data
        self.rowNumber=rowNumber
        x=self.data[0]
        y=self.data[1]
        z=self.data[2]
        w=self.data[3]
        #Data
        self.average(w,self.rowNumber)
        self.newExtremum(x,y,z,w)
        self.wNorm=self.dataProcess(w,self.rowNumber)
        self.opacity=self.gradOpacity(w,self.wMin,self.wAvr,self.wAvr,self.wMax)
        if(not(voxel3D)):
            self.plotScatterData(self.ax,x,y,z,self.wNorm,self.opacity)
        else:
            self.plotDataVoxel(self.wNorm,self.opacity)

    def setTitleLegend(self,title):
        self.ax.set_xlabel(axisName['x'])
        self.ax.set_ylabel(axisName['y'])
        self.ax.set_zlabel(axisName['z'])
        self.ax.set_title(title)

    def plotDataVoxel(self,value,opacity):
        '''
        Create cubes with voxel function of matplotlib
        value and opacity need to be list object
        '''
        #Define limit of the graphic
        lenX=int(self.xRange*2-1)
        lenY=int(self.yRange*2-1)
        lenZ=int(self.zRange*2-1)
        #Statement of arrays where coord, fill and values are stored
        fill=np.zeros((lenX,lenY,lenZ))
        newValue=np.zeros((lenX,lenY,lenZ))
        newAlpha=np.zeros((lenX,lenY,lenZ))
        x,y,z=self.data[0],self.data[1],self.data[2]
        grid_x, grid_y, grid_z = np.mgrid[self.xMin:self.xMax+self.xMinStep:self.xMinStep/2, self.yMin:self.yMax+self.yMinStep:self.yMinStep/2, self.yMin:self.zMax+self.zMinStep:self.zMinStep/2]
        #Creates little fake cube between two real cube, in the aim to see through the whole figure
        gapX=0.9*self.xMinStep
        gapY=0.9*self.yMinStep
        gapZ=0.9*self.zMinStep
        for i in range(len(grid_x)):
            if(i%2!=0):
                grid_x[i]=grid_x[i]+gapX
            for j in range(len(grid_y)):
                if(j%2!=0):
                    grid_y[i][j]=grid_y[i][j]+gapY
                for k in range(len(grid_z)):
                    if(k%2!=0):
                        grid_z[i][j][k]=grid_z[i][j][k]+gapZ
        #Fill and give a color only cubes where the point exists in the data given
        index=0
        for i in range(0,lenX):
            for j in range(0,lenY):
                for k in range(0,lenZ):
                    if(not(i%2!=0 or j%2!=0 or k%2!=0) and index<len(value)): #Coord where cube may fill
                        if((((i*self.xMinStep/2)+self.xMin)==x[index] )and ((((j*self.yMinStep/2)+self.yMin)==y[index])) and (((k*self.zMinStep/2)+self.yMin)==z[index])): #Verify if this coord exist in the file
                            if(opacity[index]!=0):
                                fill[i][j][k]=1
                                newAlpha[i][j][k]=opacity[index]
                            newValue[i][j][k]=value[index]
                            index+=1
                        else:
                            newValue[i][j][k]=self.wAvr
                            newAlpha[i][j][k]=self.wAvr
                    else:
                        pass #Already at 0
        color=colorMap(newValue,newAlpha)
        self.ax.voxels(grid_x,grid_y,grid_z,fill,facecolors=color,shade=False)
        self.draw()

class Visu3DInterp(Visu3D):
    '''
    Sub-class of Visu3D,with particular methods for interpolated data
    '''
    def __init__(self,parent=None, width=500, height=400, dpi=100):
        super().__init__(self)
        self.dataInterp=[]
        self.opacityInterp=[]
        self.gridX=[]
        self.gridY=[]
        self.gridZ=[]

    def processInterp3D(self):
        x=self.data[0]
        y=self.data[1]
        z=self.data[2]
        self.dataInterp,self.gridX,self.gridY,self.gridZ=self.interpFunc(x,y,z,self.wNorm,self.rowNumber)
        self.opacityInterp,d,d,d=self.interpFunc(x,y,z,self.opacity,self.rowNumber)
        if(not(voxel3D)):
            self.plotScatterData(self.ax,self.arrayToList(self.gridX),self.arrayToList(self.gridY),self.arrayToList(self.gridZ),self.arrayToList(self.dataInterp),alpha=self.arrayToList(self.opacityInterp))
        else:
            self.plotInterpVoxel(self.dataInterp,self.opacityInterp)

    def plotInterpVoxel(self,value,opacity):
        '''
        plotVoxel reimplemented for interpolated data
        value and opacity need to be list object
        '''
        #Define limit of the graphic
        lenX=int((self.xMax-self.xMin)/self.interpStepX*2+1)
        lenY=int((self.yMax-self.yMin)/self.interpStepY*2+1)
        lenZ=int((self.zMax-self.yMin)/self.interpStepZ*2+1)
        #Statement of arrays where coord, fill and values are stored
        fill=np.zeros((lenX,lenY,lenZ))
        newValue=np.zeros((lenX,lenY,lenZ))
        newAlpha=np.zeros((lenX,lenY,lenZ))
        grid_x, grid_y, grid_z = np.mgrid[self.xMin:self.xMax+self.interpStepX:self.interpStepX/2, self.yMin:self.yMax+self.interpStepY:self.interpStepY/2, self.yMin:self.zMax+self.interpStepZ:self.interpStepZ/2]
        #Creates little fake cube between two real cube, in the aim to see through the whole figure
        gapX=0.9*self.interpStepX
        gapY=0.9*self.interpStepY
        gapZ=0.9*self.interpStepZ
        for i in range(len(grid_x)):
            if(i%2!=0):
                grid_x[i]=grid_x[i]+gapX
            for j in range(len(grid_y)):
                if(j%2!=0):
                    grid_y[i][j]=grid_y[i][j]+gapY
                for k in range(len(grid_z)):
                    if(k%2!=0):
                        grid_z[i][j][k]=grid_z[i][j][k]+gapZ
        #Fill and give a color for real cube
        index=0
        for i in range(0,lenX):
            for j in range(0,lenY):
                for k in range(0,lenZ):
                    if(not(i%2!=0 or j%2!=0 or k%2!=0)and index<len(value)):
                        if(opacity[index]!=0):
                            fill[i][j][k]=1
                            newAlpha[i][j][k]=opacity[index]
                        newValue[i][j][k]=value[index]
                        index+=1
                    else:
                        #Already at 0
                        pass
        color=colorMap(newValue,newAlpha)
        self.ax.voxels(grid_x,grid_y,grid_z,fill,facecolors=color,shade=False)
        self.draw()


class Visu2D(Visu):
    '''
    Sub-class of visu
    '''
    def __init__(self,title, parent=None, width=500, height=400, dpi=100):
        fig=Figure(figsize=(width, height), dpi=dpi)
        super().__init__(fig)
        self.ax=self.fig.subplots()
        self.indLayer=0
        self.plane=axisName['x']+axisName['y']
        self.title=title
        self.ax.set_title(self.title)

    def process2D(self):
        '''
        Prepare data for the 2D heat map and plot it with plotMap2D method
        NB : method useless in the 1.0 version
        '''
        lenX=int(self.xRange/self.xMinStep)
        lenY=int(self.yRange/self.yMinStep)
        lenZ=int(self.zRange/self.zMinStep)
        colorMap3d=self.mat3D(self.data[0],self.data[1],self.data[2],self.wNorm,lenX,lenY,lenZ)
        self.plotMap2D(self.ax,colorMap3d,self.plane,self.indLayer,lenX,lenY,lenZ)
        self.draw()

    def plotMap2D(self,ax,colorValue,typeLayer,indLayer,lenX,lenY,lenZ,**kwargs):
        '''
        Plot 2D map at the coordinates indLayer for the plane typeLayer
        '''
        cmap=kwargs.get('cmap',colorMap)
        if(typeLayer==axisName['x']+axisName['y']):
            mat=np.zeros((lenY,lenX))
            for x in range(lenX):
                for y in range(lenY):
                    mat[y][x]=colorValue[x][y][indLayer]
            ax.set_xlabel(axisName['x'])
            ax.set_ylabel(axisName['y'])
            axisX=np.linspace(self.xMin,self.xMax,lenX)
            axisY=np.linspace(self.yMin,self.yMax,lenY)
        elif(typeLayer==axisName['x']+axisName['z']):
            mat=np.zeros((lenZ,lenX))
            for x in range(lenX):
                for z in range(lenZ):
                    mat[z][x]=colorValue[x][indLayer][z]
            ax.set_xlabel(axisName['x'])
            ax.set_ylabel(axisName['z'])
            axisX=np.linspace(self.xMin,self.xMax,lenX)
            axisY=np.linspace(self.yMin,self.zMax,lenZ)
        elif(typeLayer==axisName['y']+axisName['z']):
            mat=np.zeros((lenZ,lenY))
            for y in range(lenY):
                for z in range(lenZ):
                    mat[z][y]=colorValue[indLayer][y][z]
            ax.set_xlabel(axisName['y'])
            ax.set_ylabel(axisName['z'])
            axisX=np.linspace(self.yMin,self.yMax,lenY)
            axisY=np.linspace(self.yMin,self.zMax,lenZ)
        ax.pcolormesh(axisX,axisY,mat,cmap=cmap,shading='auto',vmin=np.nanmin(colorValue),vmax=np.nanmax(colorValue))
        self.draw()

    def setNewLayerData(self,slider,textBox):
        '''
        Change the index of the layer selected by the user in function of the plane already given
        NB : method useless in the 1.0 version
        '''
        self.indLayer=slider.value()-1
        self.ax.cla()
        self.process2D()
        if(self.plane==axisName['x']+axisName['y']):
            step=self.zMinStep
        elif(self.plane==axisName['x']+axisName['z']):
            step=self.yMinStep
        elif(self.plane==axisName['y']+axisName['z']):
            step=self.zMinStep
        textBox.setText('Index layer number '+str(slider.value()*step))


class Visu2DInterp(Visu2D):
    '''
    Sub-class of visu2D, some methods are reimplemented for interpolated data
    '''
    def __init__(self,title,parent=None, width=500, height=400, dpi=100):
        super().__init__(self,title)
        self.dataInterp=[]
        self.interpStepX,self.interpStepY,self.interpStepZ=0.5,0.5,0.5
        self.opacityInterp=[]
        self.colorBar=self.fig.colorbar(cm.ScalarMappable(cmap=colorMap))
        self.colorBar.set_ticks([0,0.5,1])

    def copy(self,visu,visuInterp):
        super().copy(visu)
        self.dataInterp,self.opacityInterp=visuInterp.dataInterp,visuInterp.opacityInterp
        self.interpStepX,self.interpStepY,self.interpStepZ=visuInterp.interpStepX,visuInterp.interpStepY,visuInterp.interpStepZ

    def plot2DInterp(self):
        '''
        (Reimplemented method)
        Prepare data for the 2D heat map and plot it with plotMap2D method
        '''
        lenX=int((self.xMax-self.xMin)/self.interpStepX+1)
        lenY=int((self.yMax-self.yMin)/self.interpStepY+1)
        lenZ=int((self.zMax-self.yMin)/self.interpStepZ+1)
        self.plotMap2D(self.ax,self.dataInterp,self.plane,self.indLayer,lenX,lenY,lenZ)
        self.colorBar.set_ticklabels([self.wMin,self.wAvr,self.wMax])
        self.draw()

    def setNewLayerInterp(self,slider,textBox):
        '''
        (Reimplemented method)
        Change the index of the layer selected by the user in function of the plane already given
        '''
        self.ax.cla()
        if(self.plane==axisName['x']+axisName['y']):
            self.indLayer=slider.value()+self.zMin-1
            step=self.interpStepZ
            self.ax.set_title(axisName['x']+' X '+axisName['y']+'\n'+axisName['z']+' '+str(slider.value()*step+self.zMin))
            textBox.setText('Index layer number '+str(slider.value()*step+self.zMin))
        elif(self.plane==axisName['x']+axisName['z']):
            self.indLayer=slider.value()+self.yMin-1
            step=self.interpStepY
            self.ax.set_title( axisName['x']+' X '+axisName['z']+'\n'+axisName['y']+' = '+str(slider.value()*step+self.yMin))
            textBox.setText('Index layer number '+str(slider.value()*step+self.yMin))
        elif(self.plane==axisName['y']+axisName['z']):
            self.indLayer=slider.value()+self.xMin-1
            step=self.interpStepX
            self.ax.set_title(axisName['y']+' X '+axisName['z']+'\n'+axisName['x']+' '+str(slider.value()*step+self.xMin))
            textBox.setText('Index layer number '+str(slider.value()*step+self.xMin))
        self.plot2DInterp()

    def startAnimation(self,slider):

        slider.setEnabled(False)
        slider.setValue(0)
        self.animation =animation.FuncAnimation(self.fig,self.mapAnimated,fargs=(slider,),frames=slider.maximum()+1,interval=animParameters['interval'],repeat=animParameters['repeat'],repeat_delay=animParameters['repeat_delay'])
        self.draw()
        slider.setEnabled(True)

    def mapAnimated(self,frames,slider):
        '''
        Anim the 2D plot map, with a scrolling of the index for plane given
        use animation module of matplotlib - Function call by the animation module
        '''
        slider.setValue(frames)

class SliderCustom(QtWidgets.QSlider):
    '''
    Slider object to control index for the heat map2D
    '''
    def __init__(self):
        super().__init__()

    def reset(self,visu,Xstep,Ystep,Zstep,textBox):
        if(visu.plane==axisName['x']+axisName['y']):
            max=(visu.zMax-visu.zMin)/Zstep
        elif(visu.plane==axisName['x']+axisName['z']):
             max=(visu.yMax-visu.yMin)/Ystep
        elif(visu.plane==axisName['y']+axisName['z']):
            max=(visu.xMax-visu.xMin)/Xstep
        self.setRange(0,int(max))
        visu.setNewLayerInterp(self,textBox)
        self.setEnabled(True)

    def setRange(self,min,max):
        self.setMinimum(min)
        self.setMaximum(max)


class OpacitySpin(QtWidgets.QDoubleSpinBox):
    '''

    '''
    def __init__(self,min,max,status,slider,window):
        super().__init__()
        self.min=min
        self.max=max
        self.setMinimum(min)
        self.setMaximum(max)
        self.w=window
        self.status=status
        self.slider=slider
        self.setPrefix(status + ' : ')
        self.valueChanged.connect(lambda x: self.newValueChanged())
        self.setStyleSheet('font-size : 14px')

    def newRange(self):
        self.disconnect()
        if(self.status=='min'):
            self.setRange(self.w.visuData3D.wMin,self.slider.value()[1])
            self.setValue(self.slider.value()[0])
        elif(self.status=='lowThreshold'):
            self.setRange(self.w.visuData3D.wMin,self.slider.value()[2])
            self.setValue(self.slider.value()[1])
        elif(self.status=='highThreshold'):
            self.setRange(self.slider.value()[1],self.w.visuData3D.wMax)
            self.setValue(self.slider.value()[2])
        elif(self.status=='max'):
            self.setRange(self.slider.value()[2],self.w.visuData3D.wMax)
            self.setValue(self.slider.value()[3])
        self.valueChanged.connect(lambda x: self.newValueChanged())

    def newValueChanged(self):

        if(self.status=='min'):
            self.w.min=int(self.value())
        elif(self.status=='lowThreshold'):
            self.w.lowThreshold=int(self.value())
        elif(self.status=='highThreshold'):
            self.w.highThreshold=int(self.value())
        elif(self.status=='max'):
            self.w.max=int(self.value())
        self.slider.setSliderPosition([self.w.min,self.w.lowThreshold,self.w.highThreshold,self.w.max])
        self.w.opacityChanged()

class MainWindow(QtWidgets.QMainWindow):

    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setWindowTitle('Data Visualize')
        #properties
        self.min=0
        self.lowThreshold=0
        self.highThreshold=0
        self.max=0

        #sheet
        sheetTitle="font-size : 22px"
        sheetLabel="font-size : 18px"
        #Canvas
        generalPanel=QtWidgets.QGridLayout()
        self.visuData3D=Visu3D(self, width=500, height=400, dpi=100)
        self.visuInterp3D=Visu3DInterp(self)
        generalPanel.addWidget(NavigationToolbar(self.visuData3D, self),0,0)
        generalPanel.addWidget(self.visuData3D,1,0)
        generalPanel.addWidget(NavigationToolbar(self.visuInterp3D, self),0,1)
        generalPanel.addWidget(self.visuInterp3D,1,1)
        self.visuInterp2D=Visu2DInterp(self,'Data Interpolated')
        generalPanel.addWidget(NavigationToolbar(self.visuInterp2D, self),2,1)
        generalPanel.addWidget(self.visuInterp2D,3,1)

        #Right Panel : Control 3D panel
        settingPanel=QtWidgets.QGridLayout()
        control3DPanel=QtWidgets.QVBoxLayout()
        #load data
        loadLayout=QtWidgets.QVBoxLayout()
        self.getDataButton=GetData('Load Data')
        self.getDataButton.setStyleSheet("background-color : darkmagenta;color:white;font-size:35px")
        self.getDataButton.clicked.connect(lambda x: self.getDataProcess())
        loadLayout.addWidget(self.getDataButton)
        control3DPanel.addLayout(loadLayout)


        #Control opacity


        self.opacityPanel=QtWidgets.QGridLayout()
        opacityTitle=QtWidgets.QLabel('Opacity settings :')
        opacityTitle.setStyleSheet(sheetTitle)
        opacityTitle.setAlignment(QtCore.Qt.AlignCenter)
        self.opacitySlider=QRangeSlider()
        self.opacitySlider.setEnabled(False)
        self.opacitySlider.sliderMoved.connect(lambda x: self.opacityChanged())
        self.opacitySlider.setOrientation(QtCore.Qt.Horizontal)
        self.opacitySpin_min=OpacitySpin(0,0,'min',self.opacitySlider,self)
        self.opacitySpin_low=OpacitySpin(0,0,'lowThreshold',self.opacitySlider,self)
        self.opacitySpin_high=OpacitySpin(0,0,'highThreshold',self.opacitySlider,self)
        self.opacitySpin_max=OpacitySpin(0,0,'max',self.opacitySlider,self)
        self.checkVoxel=QtWidgets.QCheckBox('Enable Voxel 3D\nWARNING : EXPERIMENTAL OPTION')
        self.checkVoxel.setEnabled(False)
        self.opacityPanel.addWidget(opacityTitle,0,0,1,4)
        self.opacityPanel.addWidget(self.opacitySpin_min,1,0)
        self.opacityPanel.addWidget(self.opacitySpin_low,1,1)
        self.opacityPanel.addWidget(self.opacitySpin_high,1,2)
        self.opacityPanel.addWidget(self.opacitySpin_max,1,3)
        self.opacityPanel.addWidget(self.opacitySlider,2,0,1,4)
        self.opacityPanel.addWidget(self.checkVoxel,3,0)
        control3DPanel.setSpacing(20)
        control3DPanel.addLayout(self.opacityPanel)
        control3DPanel.setAlignment(QtCore.Qt.AlignTop)
        control3DPanel.setContentsMargins(0,0,0,20)
        settingPanel.addLayout(control3DPanel,0,0)


        #2D map control
        control2DPanel=QtWidgets.QGridLayout()
        mapLabel=QtWidgets.QLabel('2D heat map settings : ')
        mapLabel.setStyleSheet(sheetTitle)
        mapLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.slider2D_interp=SliderCustom()
        self.slider2D_interp.setOrientation(QtCore.Qt.Horizontal)
        self.slider2D_interp.setEnabled(False)
        self.slider2D_interp.valueChanged.connect(lambda x: self.visuInterp2D.setNewLayerInterp(self.slider2D_interp,self.textIndLayer2D_interp))
        labelSlider2D_interp=QtWidgets.QLabel('Slider 2D interpolation')
        labelSlider2D_interp.setStyleSheet(sheetLabel)
        labelPlane=QtWidgets.QLabel('Choose plane for the 2D map :')
        labelPlane.setStyleSheet(sheetLabel)
        self.planeList=QtWidgets.QListWidget()
        self.planeList.setEnabled(False)
        self.planeList.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        self.textIndLayer2D_data=QtWidgets.QTextEdit()
        self.textIndLayer2D_interp=QtWidgets.QTextEdit()
        self.textIndLayer2D_data.setReadOnly(True)
        self.textIndLayer2D_interp.setReadOnly(True)
        self.planeList.currentItemChanged.connect(lambda x: self.newPlane())
        control2DPanel.addWidget(mapLabel,0,0,1,6)
        control2DPanel.addWidget(labelPlane,1,0,1,1)
        control2DPanel.addWidget(self.planeList,1,1,1,1)
        control2DPanel.addWidget(labelSlider2D_interp,2,0,1,4)
        control2DPanel.addWidget(self.slider2D_interp,2,1,1,4)
        control2DPanel.addWidget(self.textIndLayer2D_interp,2,5,1,1)
        control2DPanel.setColumnStretch(1,100)
        control2DPanel.setColumnStretch(5,1)
        control2DPanel.setRowStretch(1,1)
        control2DPanel.setRowStretch(0,1)
        control2DPanel.setRowStretch(2,100)

        #Record animation
        animPanel=QtWidgets.QVBoxLayout()
        self.animPush=QtWidgets.QPushButton('Save animation')
        self.animPush.setStyleSheet("background-color : darkgreen;color:white;font-size:20px")
        self.animPush.setEnabled(False)
        self.animPush.clicked.connect(lambda x:self.saveAnim())
        animPanel.addWidget(self.animPush)
        control2DPanel.addLayout(animPanel,3,1)
        control2DPanel.setAlignment(QtCore.Qt.AlignBottom)
        settingPanel.addLayout(control2DPanel,1,0)
        generalPanel.addLayout(settingPanel,3,0)

        #Panel setup
        widget=QtWidgets.QWidget()
        widget.setLayout(generalPanel)
        self.setCentralWidget(widget)
        self.showMaximized()


    def getDataProcess(self):
        '''
        Initiliaze all figures in the main window when new data are loaded
        '''
        global voxel3D
        #Clear all ploting and interface
        self.visuData3D.ax.clear()
        self.visuInterp3D.ax.clear()
        self.visuInterp2D.ax.clear()
        voxel3D=False
        self.checkVoxel.disconnect()
        self.checkVoxel.setChecked(False)
        #Get data from Excel or tsv/csv files
        data,rowNumber=self.getDataButton.loadFile()
        data,rowNumber=self.verifData(data)
        #Create legends
        self.planeList.clear()
        self.planeList.addItem(axisName['x']+axisName['y'])
        self.planeList.addItem(axisName['x']+axisName['z'])
        self.planeList.addItem(axisName['y']+axisName['z'])
        self.visuInterp2D.plane=axisName['x']+axisName['y']
        self.visuData3D.setTitleLegend('Representation of '+axisName['w']+' in function of '+axisName['x']+', '+axisName['y']+', '+axisName['z'])
        self.visuInterp3D.setTitleLegend('Representation of the interpolation of  '+axisName['w']+' in function of '+axisName['x']+', '+axisName['y']+', '+axisName['z'])
        #Set new values for atributes of visual object by affectation and copy
        self.visuData3D.process3D(data,rowNumber) #Plot raw data in 3D
        self.visuInterp3D.copy(self.visuData3D)
        self.visuInterp3D.processInterp3D() #Plot interpolate data in 3D
        self.visuInterp2D.copy(self.visuData3D,self.visuInterp3D)
        self.visuInterp2D.plot2DInterp() #Plot heat map 2D with interpolated data

        #Set new settings for the user interface
        self.setOpacityProperties(self.visuData3D.wMin,self.visuData3D.wAvr,self.visuData3D.wAvr,self.visuData3D.wMax)
        self.opacitySlider.setEnabled(True)
        self.slider2D_interp.reset(self.visuInterp2D,self.visuInterp2D.interpStepX,self.visuInterp2D.interpStepY,self.visuInterp2D.interpStepZ,self.textIndLayer2D_interp)
        self.planeList.setEnabled(True)
        self.animPush.setEnabled(True)
        self.planeList.setCurrentRow(0)
        self.checkVoxel.stateChanged.connect(lambda x: self.changeVoxel())
        self.checkVoxel.setEnabled(True)

    def setOpacityProperties(self,min,lowThreshold,highThreshold,max):
        '''
        Change the range values of opacity for the slider
        '''
        self.min=min
        self.lowThreshold=lowThreshold
        self.highThreshold=highThreshold
        self.max=max
        self.opacitySlider.setRange(self.min,self.max)
        self.opacitySlider.setValue([self.min,self.lowThreshold,self.highThreshold,self.max])
        self.opacitySpin_min.newRange()
        self.opacitySpin_low.newRange()
        self.opacitySpin_high.newRange()
        self.opacitySpin_max.newRange()

    def opacityChanged(self):
        '''
        Method connected to slider signal :
            Change opacity of the 2 figures
        '''
        #Set new threshold
        self.min=self.opacitySlider.value()[0]
        self.lowThreshold=self.opacitySlider.value()[1]
        self.highThreshold=self.opacitySlider.value()[2]
        self.max=self.opacitySlider.value()[3]
        #Compute new opacities
        self.visuData3D.opacity=self.visuData3D.gradOpacity(self.visuData3D.data[3],self.min,self.lowThreshold,self.highThreshold,self.max)
        self.visuInterp3D.opacityInterp,self.visuInterp3D.gridX,self.visuInterp3D.gridY,self.visuInterp3D.gridZ=self.visuInterp3D.interpFunc(self.visuInterp3D.data[0],self.visuInterp3D.data[1],self.visuInterp3D.data[2],self.visuData3D.opacity,self.visuInterp3D.rowNumber)
        #Plot with the new opacity
        self.updatePlot()

    def updatePlot(self):
        #clear axes
        self.visuData3D.ax.cla()
        self.visuInterp3D.ax.cla()
        #New plot
        if(not(voxel3D)):
            self.visuData3D.plotScatterData(self.visuData3D.ax,self.visuData3D.data[0],self.visuData3D.data[1],self.visuData3D.data[2],self.visuData3D.wNorm,self.visuData3D.opacity)
            self.visuInterp3D.plotScatterData(self.visuInterp3D.ax,self.visuInterp3D.arrayToList(self.visuInterp3D.gridX),self.visuInterp3D.arrayToList(self.visuInterp3D.gridY),self.visuInterp3D.arrayToList(self.visuInterp3D.gridZ),self.visuInterp3D.arrayToList(self.visuInterp3D.dataInterp),self.visuInterp3D.arrayToList(self.visuInterp3D.opacityInterp))
        else:
            self.visuData3D.plotDataVoxel(self.visuData3D.wNorm,self.visuData3D.opacity)
            self.visuInterp3D.plotInterpVoxel(self.visuInterp3D.arrayToList(self.visuInterp3D.dataInterp),self.visuInterp3D.arrayToList(self.visuInterp3D.opacityInterp))
        #Update spinbox
        self.opacitySpin_min.newRange()
        self.opacitySpin_low.newRange()
        self.opacitySpin_high.newRange()
        self.opacitySpin_max.newRange()
        #Update legend and title
        self.visuData3D.setTitleLegend('Representation of '+axisName['w']+' in function of '+axisName['x']+', '+axisName['y']+', '+axisName['z'])
        self.visuInterp3D.setTitleLegend('Representation of the interpolation of  '+axisName['w']+' in function of '+axisName['x']+', '+axisName['y']+', '+axisName['z'])
        self.visuData3D.draw()
        self.visuInterp3D.draw()

    def newPlane(self):
        '''
        Set the plane choosen by the slider for the 2D heat map
        '''
        try:
            item=self.planeList.currentItem()
            self.visuInterp2D.plane=item.text()
            self.slider2D_interp.reset(self.visuInterp2D,self.visuInterp2D.interpStepX,self.visuInterp2D.interpStepY,self.visuInterp2D.interpStepZ,self.textIndLayer2D_interp)
        except AttributeError:
            pass



    def saveAnim(self):
        filePath=QtWidgets.QFileDialog.getSaveFileName(None,'Save as...','dataHeatMap2D','*.gif *.mp4')[0]
        if(filePath !=''):
            self.visuInterp2D.startAnimation(self.slider2D_interp)
            self.visuInterp2D.animation.save(filePath, writer=animParameters['writer'], fps=animParameters['fps'])

    def verifData(self,data):
        belowOne=False
        x,y,z,value=data[0],data[1],data[2],data[3]

        #To have integer value superior to one
        for v in value:
            if v<1 and v!=0:
                belowOne=True
        if(belowOne):
            for i in range(len(value)):
                value[i]=value[i]*100
            data[3]=value
        #Verif if axes are repeated or not
        repeatedValue=False
        last_value=x[0]
        repeatCount=0
        newRowNumber=0
        for xi in x:
            if(xi>=last_value):
                if(xi>last_value):
                    last_value=xi
                if(repeatCount==0):
                    newRowNumber+=1
            elif(xi<last_value): #Detect a restart of the axes
                repeatedValue=True
                repeatCount+=1
                last_value=xi

        if(repeatedValue):
            for xi in range(newRowNumber):
                indexIncrement=0
                valueAvr=0
                for indexIncrement in range(newRowNumber,len(x)):
                    if(x[xi]==x[indexIncrement] and y[xi]==y[indexIncrement] and z[xi]==z[indexIncrement]):
                        valueAvr=value[indexIncrement]
                value[xi]=valueAvr

            #create new axes
            data=np.array((data[0][:newRowNumber],data[1][:newRowNumber],data[2][:newRowNumber],value[:newRowNumber]))
        return data,newRowNumber

    def changeVoxel(self):
        global voxel3D
        voxel3D=self.checkVoxel.isChecked()
        self.updatePlot()

#Show main window
def main():
    app = QtWidgets.QApplication(sys.argv)
    w = MainWindow()
    app.exec_()
if __name__=='__main__':
    main()